#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import os
import re
import openpyxl

excel_dict = {}
que_dict = {}
# xls = {}
# sheet = {}

txt_temp = ''

def dec_hex(str):
    pattern = r'0[xX][0-9a-fA-F]+'  # 0x 十六进制
    has_hex = re.findall(pattern, str)
    type = 0  # 判断有无单位
    unit = ''
    res = 0
    if has_hex:
        res = int(has_hex[0], 16)
    else:  # 十进制数值
        pattern1 = r'[A-Za-z]+|[-]?\d+[.]?\d*'
        has_unit = re.findall(pattern1, str)
        if len(has_unit) > 1:  # 有单位
            type = 1
            unit = has_unit[1]
            res = has_unit[0]
        elif len(has_unit) > 0:  # 无单位
            res = has_unit[0]
        else:
            res = str
    return res, unit, type  # 数值结果，单位，有无单位

def txt2excel(txt_file, name_prefix):
    abc = name_prefix.split('_')  # 获取Excel的abc列内容
    print(abc)
    if len(abc) < 4:
        print("\033[0;31m[Error] name of txt error\033[0m")
        sys.exit()
    
    pos = []  # 保存txt表头每列的起始位置
    hijks = ["TestDescription/Event", "Result", "Value", "U.limit", "L.limit", "DUT", "Pin/pattern", "Sequence"]
    head = ["LOT_ID", "WaferID", "PROBE_XY", "Temperature(C)", "Power Voltage(V)", "Pin/Pattern", "LimitUpper", "LimitLower", "Value", "Result", "Test Condition1"]

    txtlines = txt_file.readlines()  # 读取txt行数据
    for s in hijks:
        pos.append(txtlines[1].find(s))  # txt文件的第二行为表头，
    # print(pos)

    pos_len = len(pos)
    if (pos_len != len(hijks)):
        print("\033[0;31m[Error] Head num error\033[0m")
        sys.exit()
    
    pos_range = []  # 每一列的范围，按照Excel表的位置保存
    pos_range.append([pos[6], pos[7]])
    pos_range.append([pos[3], pos[4]])
    pos_range.append([pos[4], pos[5]])
    pos_range.append([pos[2], pos[3]])
    pos_range.append([pos[1], pos[2]])
    pos_range.append([pos[0], pos[1]])
    print(pos_range)

    xls = openpyxl.Workbook()

    excel_temp = ''  # 判断Sequence是否与上一条不同
    global txt_temp  # 判断是否是新的txt文件
    global excel_dict  # 字典，保存每个Excel表的行位置

    start = False

    voltage = 0.0
    for k in range(len(txtlines)): # 从头遍历txt文件

        txtline = txtlines[k]
        if txtline[0 : 6] == 'TestID':  # 判断一个txt文件有多个测试
            # k += 2
            start = True
            continue
        elif txtline[0 : 6] == '** End':
            xls.save(excel_temp + '.xlsx')
            print('[Write] save excel:', excel_temp, k)
            voltage = 0.0
            start = False
        if not start:
            continue
        # 判断此行是否为电压数值
        power = txtline[pos_range[-1][0] : pos_range[-1][1]]  
        if 'VccPower' in power:
            vcc = txtline[pos_range[3][0] : pos_range[3][1]]
            vcc = vcc.strip()
            voltage = float(vcc)
            continue
        
        excelname = txtline[pos[-1]:-1]  # 获取Sequence作为excel命名
        
        # print(excelname, k)
        if not excelname in excel_dict:  # 如果无此表，保存上一次的结果，并另新建一个
            if os.path.exists(excelname + '.xlsx'):
                print("\033[0;33m[Warn] please keep pwd no excel\033[0m")
                return
            if excel_temp != '':  # 判断是否为第一次有效行
                xls.save(excel_temp + '.xlsx')
                print('[Write] save excel:', excel_temp, k)
                xls = openpyxl.Workbook()

            print('[Write] new excel:', excelname, k)
            r = 2  # 在excel开始写的位置（y）
            c = 1  # 在excel开始写的位置（x）
            excel_dict[excelname] = r
            sheet = xls.create_sheet(excelname, 0)
            for i in range(len(head)):  # 写表头数据
                sheet.cell(row = 1, column = i + 1, value = head[i])
        else:  # 表已存在
            if txt_temp != '' and txt_temp != name_prefix:  # 如果是后续文件
                if excel_temp != '' and excel_temp != excelname:  # 非首行出现上一条不同类型，保存之前结果
                    xls.save(excel_temp + '.xlsx')
                    print('[Write] save excel:', excel_temp, k)
            else:  # 如果是第一个txt文件
                if excel_temp != excelname:
                    xls.save(excel_temp + '.xlsx')
                    print('[Write] save excel:', excel_temp, k)

            if excel_temp != excelname:  # 出现不同类型，则指向新类型
                print('[Read] load excel:', excelname, k)
                xls = openpyxl.load_workbook(excelname + '.xlsx')
                sheet = xls[excelname]
        # 上面判断主要目的指向正确的xls
        
        
        # sheet[excelname] = xls[excelname].get_sheet_by_name(excelname)
        # 对每一列写数据
        pin = txtline[pos_range[0][0] : pos_range[0][1]]
        r_f = txtline[pos_range[-2][0] : pos_range[-2][1]]
        nrow = excel_dict[excelname]
        for i in range(len(head)):
            if i < 4:
                sheet.cell(row = nrow, column = i + 1, value = abc[i])
            elif i == 4:
                sheet.cell(row = nrow, column = i + 1, value = voltage)
            elif i == 5:  # 不用考虑单位
                sheet.cell(row = nrow, column = i + 1, value = pin)
            elif 5 < i < 9:  # 考虑单位
                txt_cell = txtline[pos_range[i-5][0] : pos_range[i-5][1]]
                pattern = r'[A-Za-z]+|[-]?\d+[.]?\d*'
                result = re.findall(pattern, txt_cell)
                if len(result) > 1:
                    sheet.cell(row = 1, column = i + 1, value = head[i] + '/' + result[1])
                    sheet.cell(row = nrow, column = i + 1, value = result[0])
                elif len(result) > 0:
                    sheet.cell(row = nrow, column = i + 1, value = result[0])
                else:
                    sheet.cell(row = nrow, column = i + 1, value = txt_cell)
            elif i == 9:
                sheet.cell(row = nrow, column = i + 1, value = r_f)
            elif i == 10: # 多条件处理
                event = txtline[pos_range[-1][0] : pos_range[-1][1]]
                if len(event.split(':')) <= 1:
                    sheet.cell(row = nrow, column = i + 1, value = event.strip())
                    continue
                desc = event.split(':')[0]
                date = event.split(':')[1]  # date = ' 0x00_0x01_-0uA'
                
                descs = desc.split('_')
                dates = date.split('_')
                if (len(descs) != len(dates)):
                    print("\033[0;31m[Error] the num of events and values not same\033[0m")
                    sys.exit()
                # pattern = r'(.*?);'
                # result = re.findall(pattern, date) # result = {' 0x00', ' 0x01', ' -0uA'}
                for j in range(len(dates)):
                    decimal, unit, type = dec_hex(dates[j])
                    if type:
                        test = descs[j] + '/' + unit
                    else:
                        test = descs[j]
                    sheet.cell(row = nrow, column = j + i + 1, value = decimal)
                    sheet.cell(row = 1, column = j + i + 1, value = test)
                        
        excel_dict[excelname] += 1
        excel_temp = excelname
    txt_temp = name_prefix

def fun1(xl):
    xls = openpyxl.load_workbook(xl)
    source = xls[xl.split('.')[0]]  # sheet
    cc = 9
    unit = ''
    pre = ''
    while True:  # 找到Dc的列，赋值给cc，如果没有则下一个文件
        cell_temp = source.cell(row=1, column=cc)
        if cell_temp.value:
            temp = cell_temp.value.split('/')[0]
            if temp == 'Dc' or temp == 'DC':
                pre = 'Dc-'
                unit = cell_temp.value.split('/')[1]
                break
            elif temp == 'T':
                pre = 'T-'
                break
            else:
                cc += 1
        else:
            return
    print('\033[0;32mStart function in xlsx: %s\033[0m' % xl)
    print(cc)
    target = xls.create_sheet('test', 1)
    t_cc_s = cc + 1
    t_cc = t_cc_s  # t_cc代表目标sheet的列位置
    t_rr = 2  # 同上
    index_c = cc - 1  # list遍历时用到的指针
    head_num = []  # Dc对应的值 0 10 20 30 0 10 20
    head_str = []  # T对应的字符串
    type1 = 0  # T or Dc
    val_pre = 0  # 用来判断换行
    once = True
    m_r = source.max_row
    print(m_r)
    list_line = list(source.values)  # 读取sheet的全部值
    
    for index_r in range(1, m_r):
        cell_str = source.cell(row=index_r+2, column=cc).value
        if not cell_str:  # 存在没有值的情况
            continue
        if once:
            once = False
            cell_str = str(cell_str)
            pattern = r'[0-9]+'  # 数字
            has_num = re.findall(pattern, cell_str)
            if has_num:
                if len(cell_str.split('.')) > 1:
                    type1 = 1  # 小数
                else:
                    type1 = 2  # 整数
            else:
                type1 = 3  # 字符串
            target.append(list_line[0])  # 头目录
        line = list_line[index_r]  # 遍历原sheet的行
        cell_str = source.cell(row=index_r+1, column=9).value
        if '----' in cell_str:
            target.append(line)
            t_rr += 1
            continue
        if type1 == 1:
            val = float(line[index_c])
        elif type1 == 2:
            val = int(line[index_c])
        elif type1 == 3:
            strr = line[index_c]
        else:
            print("\033[0;31m[Error] in 227 line\033[0m")
        # print(type(line_r))
        # target.append(list_line[i])
        if type1 == 3:  # T处理仅仅横排显示
            # if len(head_str) == 0:
            #     head_str.append(strr)
            #     t_cc += 1
            #     # target.append(list_line[0])
                
            #     target.cell(row=1, column=t_cc, value=pre + str(strr))
            if len(head_str) == 0 or head_str[-1] != strr:
                head_str.append(strr)
                t_cc += 1
                target.cell(row=1, column=t_cc, value=pre + str(strr))
            else:
                t_cc += 1
                print('\033[0;33m[Warn]T do nothing\033[0m')
            
            target.append(line)
            target.cell(row=t_rr, column=t_cc, value=line[8])
        else:
            if len(head_num) == 0:  # 第一行
                head_num.append(val)
                t_cc += 1
                
                target.append(line)  # 第一行内容
                target.cell(row=1, column=t_cc, value=pre + str(val) + unit)  # 通过t_cc确定横排的位置
                # target.cell(row=2, column=t_cc, value=source.cell(row=rr, column=8).value)
            elif head_num[-1] < val:  # 如果Dc的值出现更大的情况
                head_num.append(val)
                t_cc += 1
                target.cell(row=1, column=t_cc, value=pre + str(val) + unit)
            else:
                t_cc = t_cc_s
                if val < val_pre:  # 只在数据下降沿换行
                    t_rr += 1
                    target.append(line)
                for i in range(len(head_num)):
                    if head_num[i] == val:
                        t_cc += i + 1
            # print('[%d, %d]' % (t_rr, t_cc))
            val_pre = val
            target.cell(row=t_rr, column=t_cc, value=line[8])  # 填充value值
    xls.remove(source)
    target.title = xl.split('.')[0]
    xls.save(xl)

def merger_excel(xl_file):
    # IO1V8_XX IO3V3_XX前缀相同的合并到一个excel的不同sheet中
    prefix = {}  # 字典，文件名前缀出现次数
    key = []  # 列表，保存需要合并的文件的关键字
    filelist = []  # 保存需要合并的文件

    for file in xl_file:
        name = file.split('.xlsx')[0]
        name = name.split('_')[0]
        if not name in prefix:
            prefix[name] = 1
        else:
            prefix[name] += 1
    for k in prefix.keys():
        if prefix[k] > 1:
            key.append(k)
    print(key)
    
    for i in range(len(key)):
        filelist_temp = []
        for file in xl_file:
            if key[i] == file.split('_')[0]:
                filelist_temp.append(file)
        filelist.append(filelist_temp)
    print(filelist)  # 需要合并的文件列表

    for i in range(len(filelist)):
        if len(filelist[i]) == 1:
            print("\033[0;33m[Warn] only one excel]\033[0m")
            continue
        src = openpyxl.load_workbook(filelist[i][0])
        print('[Read] open ' + filelist[i][0])
        # 取第一个文件，将剩下的文件复制到第一个文件的sheet
        for j in range(1, len(filelist[i])):
            target = openpyxl.load_workbook(filelist[i][j]).active
            target._parent = src
            src._add_sheet(target)
            sheet_name = filelist[i][j].split('.')[0]
            # target_sheet = target.get_sheet_by_name(sheet_name)
            # src.copy_worksheet(target_sheet)
            print('[Merger] sheet ' + sheet_name)
            os.remove(filelist[i][j])
            print('[Delete] sheet ' + sheet_name)
        ws = src['Sheet']
        src.remove(ws)
        src.save(filelist[i][0])
        test_name = ''
        #重命名处理
        if key[i] == 'IO1V8' or key[i] == 'IO3V3':
            test_name = '_DC_TEST'
        else:
            test_name = '_TEST'
        os.rename(filelist[i][0], key[i] + test_name + '.xlsx')
        print('[Name] new excel name ' + key[i] + 'DC_TEST' + '.xlsx')


if __name__ == "__main__":
    
    txt_num = len(sys.argv)
    txt_prefix = []  # 保存txt文件前缀名
    xl_file = []
    merger = True
    fun = True
    fun_key = ["VCCIO", "HVPP", "VLD"]
    fun_list = []
    for i in range(1, txt_num):
        if os.path.exists(sys.argv[i]):  # 如果文件名存在
            txt_prefix.append(os.path.basename(sys.argv[i]).split('.txt')[0])
            print('[Read] txtFile %s : %s' % (i, sys.argv[i]))
        else:
            print("\033[0;31m[Error] txtFile %s : %s not found\033[0m" % (i, sys.argv[i]))
            sys.exit()
    
    for i in range(1, txt_num):
        txtfile = open(sys.argv[i], 'r')
        print('\033[0;32mStart transform txt: %s\033[0m' % sys.argv[i])
        txt2excel(txtfile, txt_prefix[i - 1])
        txtfile.close()
    
    listd = os.listdir('./')  # 文件列表
    for file in listd:
        if file.endswith('.xlsx'):  # excel文件
            xl_file.append(file)
    
    for key in fun_key:  # 对特定文件判断
        for file in xl_file:
            if key in file.split('_')[0]:
                fun_list.append(file)
    if fun:
        for xl in fun_list:
            fun1(xl)
    
    if merger:  # 合并文件
        print('\033[0;32mStart merger excel...\033[0m')
        merger_excel(xl_file)
